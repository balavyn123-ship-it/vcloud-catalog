"""
Крок 1: Переносить всі товари з products.js → catalog.xlsx
Крок 2: Додає колонки: URL, Статус, AI перевірка
"""
import re
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

JS_FILE    = "../js/products.js"
EXCEL_FILE = "catalog.xlsx"

# ─── Читаємо products.js ───────────────────────────────────────
print("📖 Читаю products.js...")
with open(JS_FILE, "r", encoding="utf-8") as f:
    content = f.read()

# Витягуємо масив товарів
match = re.search(r"const products = (\[[\s\S]*\]);", content)
if not match:
    print("❌ Не знайшов масив products у файлі!")
    exit()

products = json.loads(match.group(1))
print(f"✅ Знайдено {len(products)} товарів")

# ─── Створюємо Excel ───────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Каталог VclouD"

# Стилі
DARK  = "1a1a2e"
GREEN = "27ae60"
GRAY  = "95a5a6"
WHITE = "FFFFFF"

def header_cell(cell, text, bg=DARK):
    cell.value = text
    cell.font = Font(bold=True, color=WHITE, size=11)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

# ─── Заголовки ────────────────────────────────────────────────
headers = [
    ("A", "ID",           8),
    ("B", "Назва товару", 45),
    ("C", "Ціна (грн)",   12),
    ("D", "Категорія",    15),
    ("E", "URL для парсингу\n(вставляй посилання сюди)", 45),
    ("F", "Поточне фото", 40),
    ("G", "Статус",       15),
    ("H", "AI перевірка", 30),
]

for col_letter, title, width in headers:
    col_idx = ord(col_letter) - ord("A") + 1
    cell = ws.cell(row=1, column=col_idx)
    header_cell(cell, title)
    ws.column_dimensions[col_letter].width = width

ws.row_dimensions[1].height = 40
ws.freeze_panes = "A2"

# ─── Заповнюємо товарами ──────────────────────────────────────
status_colors = {
    "має фото":    "d5f5e3",
    "без фото":    "fadbd8",
    "потребує URL": "fef9e7",
}

for i, p in enumerate(products, 2):
    has_image = bool(p.get("image") or p.get("images"))
    has_url   = bool(p.get("olxUrl") or p.get("sourceUrl"))

    # Статус
    if has_image:
        status = "має фото ✅"
        row_color = "f9f9f9" if i % 2 == 0 else "ffffff"
    else:
        status = "без фото ❌"
        row_color = "fadbd8"

    url = p.get("olxUrl") or p.get("sourceUrl") or ""
    image_url = p.get("image") or (p.get("images", [None])[0] or "")

    cells = [
        (1, p.get("id", i-1)),
        (2, p.get("title", "")),
        (3, p.get("price", 0)),
        (4, p.get("category", "")),
        (5, url),
        (6, image_url),
        (7, status),
        (8, ""),   # AI перевірка — заповниться після парсингу
    ]

    for col, value in cells:
        cell = ws.cell(row=i, column=col, value=value)
        cell.alignment = Alignment(vertical="center", wrap_text=False)
        cell.fill = PatternFill("solid", fgColor=row_color)
        cell.border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"),  bottom=Side(style="thin")
        )

    # Виділяємо ціну жирним
    ws.cell(row=i, column=3).font = Font(bold=True)

    # Виділяємо статус кольором
    status_cell = ws.cell(row=i, column=7)
    if "✅" in status:
        status_cell.fill = PatternFill("solid", fgColor="d5f5e3")
        status_cell.font = Font(color="1e8449", bold=True)
    else:
        status_cell.fill = PatternFill("solid", fgColor="fadbd8")
        status_cell.font = Font(color="c0392b", bold=True)

# ─── Легенда ──────────────────────────────────────────────────
ws2 = wb.create_sheet("📋 Інструкція")
instructions = [
    ("Як додати нові товари:", None, True),
    ("", None, False),
    ("1. В колонці B впиши назву товару", None, False),
    ("2. В колонці C впиши ціну в гривнях", None, False),
    ("3. В колонці E встав посилання на товар:", None, False),
    ("   • Rozetka:  https://rozetka.com.ua/...", None, False),
    ("   • Prom.ua:  https://prom.ua/...", None, False),
    ("   • Amazon:   https://amazon.com/...", None, False),
    ("   • OLX:      https://olx.ua/...", None, False),
    ("   • Будь-який інший сайт з товаром", None, False),
    ("", None, False),
    ("4. Збережи файл (Cmd+S)", None, False),
    ("5. Запусти: ./deploy.sh", None, False),
    ("", None, False),
    ("Результат:", None, True),
    ("   ✅ Парсер відкриє саме твоє посилання", None, False),
    ("   ✅ Візьме фото з тієї сторінки", None, False),
    ("   ✅ Візьме опис з тієї сторінки", None, False),
    ("   ✅ AI перевірить чи правильно спарсилось", None, False),
    ("   ✅ Сайт оновиться автоматично", None, False),
]

ws2.column_dimensions["A"].width = 60
for r, (text, _, bold) in enumerate(instructions, 1):
    cell = ws2.cell(row=r, column=1, value=text)
    cell.font = Font(bold=bold, size=12)
    ws2.row_dimensions[r].height = 20

wb.save(EXCEL_FILE)

# Статистика
with_photo    = sum(1 for p in products if p.get("image") or p.get("images"))
without_photo = len(products) - with_photo
with_url      = sum(1 for p in products if p.get("olxUrl") or p.get("sourceUrl"))

print(f"\n✅ Збережено в {EXCEL_FILE}")
print(f"   Всього товарів:   {len(products)}")
print(f"   З фото:           {with_photo}")
print(f"   Без фото:         {without_photo}")
print(f"   З посиланням:     {with_url}")
print(f"\n💡 Для товарів без фото — встав URL в колонку E і запусти парсер!")
