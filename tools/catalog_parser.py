"""
VclouD Smart Catalog Parser v2
================================
Читає catalog.xlsx:
  - Якщо є URL (колонка E) → парсить САМЕ з того посилання
  - Якщо немає URL → шукає за назвою (Rozetka/Prom)
  - AI перевірка: порівнює назву з тим що знайшло
  - Оновлює products.js і позначає статус в Excel
"""

import openpyxl
import requests
import json
import re
import os
import time
from bs4 import BeautifulSoup
from datetime import datetime
from urllib.parse import quote_plus, urlparse
from openpyxl.styles import Font, PatternFill

# ─── Налаштування ──────────────────────────────────────────────
EXCEL_FILE = "catalog.xlsx"
OUTPUT_JS  = "../js/products.js"
DELAY      = 2

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "uk-UA,uk;q=0.9,en;q=0.8",
}

CATEGORY_KEYWORDS = {
    "clothing":    ["куртка","футболка","джинси","штани","сукня","кофта","светр",
                    "пальто","шорти","сорочка","блуза","жилет","nike","adidas","zara","h&m"],
    "electronics": ["iphone","samsung","ноутбук","телефон","смартфон","планшет",
                    "навушники","телевізор","принтер","камера","xiaomi","apple","sony","lg","hp"],
    "accessories": ["сумка","рюкзак","гаманець","пояс","окуляри","годинник","прикраса","браслет"],
    "shoes":       ["кросівки","черевики","туфлі","сандалі","кеди","чоботи","мокасини"],
}


# ─── Визначення категорії ───────────────────────────────────────
def detect_category(title):
    t = title.lower()
    for cat, kws in CATEGORY_KEYWORDS.items():
        if any(k in t for k in kws):
            return cat
    return "other"


# ─── AI перевірка (локальна, без API ключа) ─────────────────────
def ai_check(search_title, found_title, found_images):
    """Перевіряє чи знайдений товар відповідає запиту."""
    if not found_title:
        return "❌ не знайдено", "Товар не знайдено на сторінці"

    search_words = set(re.findall(r'\b\w{3,}\b', search_title.lower()))
    found_words  = set(re.findall(r'\b\w{3,}\b', found_title.lower()))
    common = search_words & found_words
    score  = len(common) / max(len(search_words), 1)

    if not found_images:
        return "⚠️ без фото", f"Знайдено «{found_title[:40]}», але немає фото"
    if score >= 0.5:
        return "✅ правильно", f"Знайдено «{found_title[:50]}»"
    elif score >= 0.2:
        return "⚠️ можливо", f"Схожий: «{found_title[:50]}»"
    else:
        return "❌ інший товар", f"Знайдено «{found_title[:50]}» — не схоже на запит"


# ─── Парсинг конкретного URL ─────────────────────────────────────
def parse_from_url(url):
    """Парсить САМЕ ту сторінку яку вказав користувач"""
    domain = urlparse(url).netloc.lower()
    print(f"  🌐 Парсю: {domain}")
    try:
        resp = requests.get(url, headers=HEADERS, timeout=15)
        soup = BeautifulSoup(resp.text, "html.parser")
        images, description, found_title = [], "", ""

        if "rozetka" in domain:
            h1 = soup.find("h1", class_=re.compile("product__title|title__product"))
            found_title = h1.get_text(strip=True) if h1 else ""
            desc_div = soup.find("div", class_=re.compile("product-description|description"))
            description = desc_div.get_text(strip=True)[:500] if desc_div else ""
            for meta in soup.find_all("meta", property="og:image"):
                img = meta.get("content", "")
                if img and img not in images:
                    images.append(img)

        elif "prom.ua" in domain:
            h1 = soup.find("h1")
            found_title = h1.get_text(strip=True) if h1 else ""
            desc_div = soup.find("div", class_=re.compile("description|about"))
            description = desc_div.get_text(strip=True)[:500] if desc_div else ""
            for meta in soup.find_all("meta", property="og:image"):
                img = meta.get("content", "")
                if img and img not in images:
                    images.append(img)

        elif "olx" in domain:
            h1 = soup.find("h1")
            found_title = h1.get_text(strip=True) if h1 else ""
            # OLX: беремо дані з og: мета-тегів (вони завжди є)
            og_title = soup.find("meta", property="og:title")
            og_desc  = soup.find("meta", property="og:description")
            if og_title and not found_title:
                found_title = og_title.get("content", "")
            description = og_desc.get("content", "")[:500] if og_desc else ""
            for meta in soup.find_all("meta", property="og:image"):
                img = meta.get("content", "")
                if img and img not in images and "placeholder" not in img:
                    images.append(img)

        elif "amazon" in domain:
            h1 = soup.find("span", id="productTitle")
            found_title = h1.get_text(strip=True) if h1 else ""
            desc_div = soup.find("div", id="productDescription")
            description = desc_div.get_text(strip=True)[:500] if desc_div else ""
            img_block = soup.find("img", id="landingImage")
            if img_block:
                src = img_block.get("src") or img_block.get("data-src", "")
                if src:
                    images.append(src)

        else:
            h1 = soup.find("h1")
            found_title = h1.get_text(strip=True) if h1 else ""
            for meta in soup.find_all("meta", property="og:image"):
                img = meta.get("content", "")
                if img and img not in images:
                    images.append(img)
            desc_meta = soup.find("meta", {"name": "description"})
            description = desc_meta.get("content", "") if desc_meta else ""

        # Якщо фото не знайшли через HTML — беремо з DuckDuckGo
        if not images:
            print(f"  🔄 HTML не дав фото → шукаю через DuckDuckGo...")
            images = search_images_ddg(found_title or "product")

        return {
            "found_title": found_title,
            "description": description,
            "images":      images[:5],
            "image":       images[0] if images else "",
            "source":      domain,
        }
    except Exception as e:
        print(f"  ❌ Помилка: {e}")
        return None


# ─── Пошук фото через Bing Images ──────────────────────────────
def search_images_ddg(title, count=5):
    """Шукає фото товару через Bing Images"""
    try:
        query = quote_plus(f"{title} товар купити")
        r = requests.get(
            f"https://www.bing.com/images/search?q={query}&form=HDRSC2&first=1",
            headers=HEADERS, timeout=12
        )
        # Bing вбудовує URL фото в HTML у двох форматах
        matches = re.findall(r'murl&quot;:&quot;(https[^&]+)&quot;', r.text)
        if not matches:
            matches = re.findall(r'"murl":"(https[^"]+)"', r.text)

        # Беремо тільки реальні фото (jpg/png/webp)
        images = []
        for url in matches:
            if any(ext in url.lower() for ext in ['.jpg', '.jpeg', '.png', '.webp']):
                if url not in images:
                    images.append(url)
            if len(images) >= count:
                break

        # Якщо немає з розширенням — беремо будь-які
        if not images:
            images = list(dict.fromkeys(matches))[:count]

        return images
    except Exception as e:
        print(f"  ⚠️ Bing Images: {e}")
        return []


# ─── Пошук назви/опису на Rozetka через API ─────────────────────
def search_rozetka(title):
    try:
        url = f"https://search.rozetka.com.ua/ua/search/api/v6/?front-type=xl&lang=ua&text={quote_plus(title)}"
        resp = requests.get(url, headers=HEADERS, timeout=10)
        data = resp.json()
        goods = data.get("data", {}).get("goods", [])
        if not goods:
            return None

        # Беремо назву з API, фото — з DuckDuckGo
        item = goods[0]
        found_title = item.get("title") or title
        images = search_images_ddg(found_title or title)

        return {
            "found_title": found_title,
            "description": found_title,
            "images":      images,
            "image":       images[0] if images else "",
            "source":      "rozetka+ddg",
        }
    except Exception as e:
        print(f"  ⚠️ Rozetka: {e}")
        # Якщо Rozetka впала — просто шукаємо фото
        images = search_images_ddg(title)
        return {
            "found_title": title,
            "description": title,
            "images":      images,
            "image":       images[0] if images else "",
            "source":      "ddg-only",
        } if images else None


# ─── Читання Excel ──────────────────────────────────────────────
def read_excel():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    items = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), 2):
        title = row[1].value  # Колонка B
        price = row[2].value  # Колонка C
        url   = row[4].value  # Колонка E
        if not title or not price:
            continue
        items.append({
            "row":   row_idx,
            "title": str(title).strip(),
            "price": int(float(str(price).replace(",", "."))),
            "url":   str(url).strip() if url and str(url).strip() not in ("None", "") else None,
        })
    return wb, ws, items


# ─── Читання існуючих товарів ────────────────────────────────────
def load_existing_products():
    if not os.path.exists(OUTPUT_JS):
        return [], 0
    with open(OUTPUT_JS, "r", encoding="utf-8") as f:
        content = f.read()
    match = re.search(r"const products = (\[[\s\S]*?\]);", content)
    if match:
        try:
            products = json.loads(match.group(1))
            return products, max((p.get("id", 0) for p in products), default=0)
        except:
            pass
    return [], 0


# ─── Збереження products.js ──────────────────────────────────────
def save_products(products):
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    js = f"""// VclouD Products - Auto-generated
// Generated: {now}
// Total products: {len(products)}

const products = {json.dumps(products, ensure_ascii=False, indent=2)};
"""
    with open(OUTPUT_JS, "w", encoding="utf-8") as f:
        f.write(js)
    print(f"\n✅ Збережено {len(products)} товарів → {OUTPUT_JS}")


# ─── Оновлення статусу в Excel ───────────────────────────────────
def update_excel_status(ws, row_idx, status, ai_comment, image_url):
    ws.cell(row=row_idx, column=6).value = image_url
    status_cell = ws.cell(row=row_idx, column=7)
    status_cell.value = status
    if "✅" in status:
        status_cell.fill = PatternFill("solid", fgColor="d5f5e3")
        status_cell.font = Font(color="1e8449", bold=True)
    elif "⚠️" in status:
        status_cell.fill = PatternFill("solid", fgColor="fef9e7")
        status_cell.font = Font(color="b7950b", bold=True)
    else:
        status_cell.fill = PatternFill("solid", fgColor="fadbd8")
        status_cell.font = Font(color="c0392b", bold=True)
    ws.cell(row=row_idx, column=8).value = ai_comment


# ─── ГОЛОВНА ФУНКЦІЯ ─────────────────────────────────────────────
def main():
    print("=" * 55)
    print("  VclouD Smart Catalog Parser v2")
    print("=" * 55)

    if not os.path.exists(EXCEL_FILE):
        print(f"❌ Не знайдено {EXCEL_FILE}")
        return

    wb, ws, items = read_excel()
    existing_products, last_id = load_existing_products()
    existing_titles = {p["title"].lower(): p for p in existing_products}

    print(f"📊 В Excel:     {len(items)} рядків")
    print(f"📦 На сайті:    {len(existing_products)} товарів")

    new_items = [i for i in items if i["title"].lower() not in existing_titles]
    print(f"🆕 Нових:       {len(new_items)}")

    if not new_items:
        print("\n✅ Нових товарів немає. Всі вже є на сайті.")
        return

    added = []
    for item in new_items:
        title, price, url, row = item["title"], item["price"], item["url"], item["row"]
        print(f"\n{'─'*50}")
        print(f"🛒  {title} | {price} грн")

        if url:
            print(f"  🔗 URL знайдено → парсю з посилання")
            data = parse_from_url(url)
        else:
            print(f"  🔍 URL немає → шукаю на Rozetka")
            data = search_rozetka(title)

        if data:
            ai_status, ai_comment = ai_check(title, data.get("found_title", ""), data.get("images", []))
        else:
            ai_status, ai_comment = "❌ не знайдено", "Нічого не знайдено"

        print(f"  🤖 AI: {ai_status} — {ai_comment}")

        last_id += 1
        image_url = data.get("image", "") if data else ""
        product = {
            "id":          last_id,
            "title":       title,
            "price":       price,
            "currency":    "UAH",
            "category":    detect_category(title),
            "description": data.get("description", title) if data else title,
            "images":      data.get("images", []) if data else [],
            "image":       image_url,
            "sourceUrl":   url or "",
            "aiCheck":     ai_status,
            "date":        datetime.now().strftime("%Y-%m-%d"),
        }
        added.append(product)
        update_excel_status(ws, row, ai_status, ai_comment, image_url)
        time.sleep(DELAY)

    all_products = existing_products + added
    save_products(all_products)
    wb.save(EXCEL_FILE)

    print(f"\n📊 Excel оновлено з AI статусами")
    print(f"\n{'='*55}")
    print(f"🎉 Готово! Додано: {len(added)} нових товарів")
    print(f"📊 Всього на сайті: {len(all_products)}")
    print(f"➡️  Запусти ./deploy.sh щоб оновити сайт")
    print(f"{'='*55}")


if __name__ == "__main__":
    main()
