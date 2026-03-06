"""
Створює шаблон catalog.xlsx з правильними колонками
Запусти один раз: python3 create_template.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Каталог"

# Стилі заголовків
header_font = Font(bold=True, color="FFFFFF", size=12)
header_fill = PatternFill("solid", fgColor="1a1a2e")
center = Alignment(horizontal="center", vertical="center")
border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Заголовки
headers = ["Назва товару", "Ціна (грн)"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border

# Ширина колонок
ws.column_dimensions["A"].width = 45
ws.column_dimensions["B"].width = 15
ws.row_dimensions[1].height = 30

# Приклади (можна видалити)
examples = [
    ("Nike Air Max 90 чорні", 2500),
    ("Куртка зимова Columbia чоловіча", 3200),
    ("iPhone 13 Pro 256GB", 18000),
    ("Джинси Levis 501 сині", 1800),
    ("Samsung Galaxy Watch 5", 5500),
]

example_font = Font(color="888888", italic=True)
for row_idx, (name, price) in enumerate(examples, 2):
    ws.cell(row=row_idx, column=1, value=name).font = example_font
    ws.cell(row=row_idx, column=2, value=price).font = example_font

# Захист заголовка
ws.freeze_panes = "A2"

wb.save("catalog.xlsx")
print("✅ Створено catalog.xlsx")
print("   Заповни своїми товарами і запусти: python3 catalog_parser.py")
